from pathlib import Path

from openpyxl import load_workbook

from scripts.build_refreshed_xlsx import (
    build,
    build_pivot_cells,
    build_role_catalog,
    build_vendor_cert_map,
    read_v21_certification_rows,
)

FIXTURE = Path(__file__).parent / "fixtures" / "tiny_v21.xlsx"


def test_read_rows_returns_dicts_with_expected_keys():
    rows = read_v21_certification_rows(FIXTURE)
    assert len(rows) == 9  # matches fixture data
    expected_keys = {"wrc", "work_role_name", "element", "acronym", "proficiency", "vendor"}
    assert set(rows[0].keys()) == expected_keys


def test_role_catalog_dedupes_and_strips():
    rows = read_v21_certification_rows(FIXTURE)
    catalog = build_role_catalog(rows)
    assert set(catalog.keys()) == {"111", "211", "411"}
    # Trailing spaces in fixture's 'All-Source Analyst ' should be stripped
    assert catalog["111"]["name"] == "All-Source Analyst"
    assert catalog["111"]["element"] == "Intelligence (Cyberspace)"


def test_vendor_cert_map_groups_certs():
    rows = read_v21_certification_rows(FIXTURE)
    vmap = build_vendor_cert_map(rows)
    # Fixture vendors: Rocheston LLC, CompTIA Inc., EC-Council, GIAC
    assert "CompTIA, Inc." in vmap
    # CompTIA certs sorted: A+, CySA+, Network+, PenTest+, Security+
    assert vmap["CompTIA, Inc."] == sorted(set(vmap["CompTIA, Inc."]))
    assert "A+" in vmap["CompTIA, Inc."]
    assert "Network+" in vmap["CompTIA, Inc."]


def test_pivot_cells_basic():
    rows = read_v21_certification_rows(FIXTURE)
    cells = build_pivot_cells(rows)
    # Fixture: (411, A+, Basic=1), (411, Network+, Basic=1), (411, Security+, Intermediate=2)
    assert cells[("411", "A+")] == 1
    assert cells[("411", "Network+")] == 1
    assert cells[("411", "Security+")] == 2
    # Fixture: (211, CHFI, Intermediate=2), (211, PenTest+, Intermediate=2),
    #         (211, CySA+, Advanced=3), (211, GCFA, Advanced=3)
    assert cells[("211", "CHFI")] == 2
    assert cells[("211", "CySA+")] == 3


def test_pivot_cells_uses_highest_level_when_cert_appears_at_multiple_levels():
    # Manually craft rows where same (role, cert) appears at Basic AND Advanced
    rows = [
        {"wrc": "111", "work_role_name": "X", "element": "E",
         "acronym": "Security+", "proficiency": "basic", "vendor": "CompTIA, Inc."},
        {"wrc": "111", "work_role_name": "X", "element": "E",
         "acronym": "Security+", "proficiency": "advanced", "vendor": "CompTIA, Inc."},
    ]
    cells = build_pivot_cells(rows)
    assert cells[("111", "Security+")] == 3  # advanced wins


def test_build_produces_three_sheets(tmp_path):
    out = tmp_path / "out.xlsx"
    build(FIXTURE, out)
    wb = load_workbook(out, read_only=True)
    assert wb.sheetnames == [
        "Explanation",
        "Certification Requirements",
        "Certification Analysis",
    ]


def test_summary_sheet_has_one_row_per_role(tmp_path):
    out = tmp_path / "out.xlsx"
    build(FIXTURE, out)
    wb = load_workbook(out, data_only=True)
    ws = wb["Certification Requirements"]
    # Header rows + 3 data rows for 3 fixture roles
    role_codes = []
    for row in ws.iter_rows(min_row=3, values_only=True):
        if row[0]:
            role_codes.append(str(row[0]))
    # Role cell formatted as "(CODE) Name" per template convention
    assert any("411" in r for r in role_codes)
    assert any("211" in r for r in role_codes)
    assert any("111" in r for r in role_codes)


def test_pivot_sheet_has_certs_as_columns_and_roles_as_rows(tmp_path):
    out = tmp_path / "out.xlsx"
    build(FIXTURE, out)
    wb = load_workbook(out, data_only=True)
    ws = wb["Certification Analysis"]

    # Find the header row with cert acronyms (should be row 3 per template)
    header_row = [c.value for c in ws[3]]
    # Work Role column first, then cert columns, short names per visual_spec
    assert header_row[0] and "Work Role" in str(header_row[0])
    assert "A+" in header_row
    # Security+ gets shortened to "Sec+" per visual_spec.CERT_SHORT_NAMES
    assert "Sec+" in header_row
    assert "Net+" in header_row  # Network+ -> Net+

    # Find the row for role (411) and check A+ cell = 1
    for row in ws.iter_rows(min_row=4):
        if row[0].value and "411" in str(row[0].value):
            a_plus_col = header_row.index("A+")
            assert row[a_plus_col].value == 1
            break
    else:
        assert False, "role 411 row not found in pivot sheet"


def test_pivot_sheet_has_summary_rows(tmp_path):
    out = tmp_path / "out.xlsx"
    build(FIXTURE, out)
    wb = load_workbook(out, data_only=True)
    ws = wb["Certification Analysis"]
    all_labels = [str(r[0].value or "") for r in ws.iter_rows()]
    joined = " | ".join(all_labels)
    assert "Total Positions Covered" in joined
    assert "Points" in joined  # Label is 'Total "Points" ...' (quoted)


def test_pivot_sheet_summary_uses_formulas_not_hardcoded(tmp_path):
    # Skill requirement: summary cells must be Excel formulas, not Python-computed values.
    out = tmp_path / "out.xlsx"
    build(FIXTURE, out)
    wb = load_workbook(out)  # data_only=False so we see formulas
    ws = wb["Certification Analysis"]
    # Find the Total Positions row
    totals_row_num = None
    for row in ws.iter_rows():
        if row[0].value and "Total Positions Covered" in str(row[0].value):
            totals_row_num = row[0].row
            break
    assert totals_row_num is not None
    # Column B should hold a formula like =COUNT(B4:B...)
    formula_cell = ws.cell(row=totals_row_num, column=2)
    assert isinstance(formula_cell.value, str)
    assert formula_cell.value.startswith("=COUNT(")


def test_pivot_sheet_has_echo_column_on_right(tmp_path):
    # Work Role label column duplicated on the right side
    out = tmp_path / "out.xlsx"
    build(FIXTURE, out)
    wb = load_workbook(out, data_only=True)
    ws = wb["Certification Analysis"]
    last_col = ws.max_column
    assert ws.cell(row=3, column=last_col).value == "Work Role"
    # First data row's echo cell should repeat the left-column label
    left = ws.cell(row=4, column=1).value
    right = ws.cell(row=4, column=last_col).value
    assert left == right


def test_pivot_sheet_cert_headers_are_rotated(tmp_path):
    # Per Jan 2025 styling: cert acronym header cells are rotated 90 degrees.
    out = tmp_path / "out.xlsx"
    build(FIXTURE, out)
    wb = load_workbook(out)
    ws = wb["Certification Analysis"]
    # Column B (first cert) in row 3 should be rotated
    cell = ws.cell(row=3, column=2)
    assert cell.alignment.text_rotation == 90
