"""Build a tiny fixture xlsx mimicking the Jan 2025 'Education or Certification' sheet.

Run once from repo root:
    .venv/Scripts/python -m scripts.tests.build_jan2025_fixture
"""
from pathlib import Path
from openpyxl import Workbook


def build() -> None:
    wb = Workbook()
    wb.remove(wb.active)

    exp = wb.create_sheet("Explanation")
    exp["A1"] = "DoDM 8140.03 test fixture"

    ws = wb.create_sheet("Education or Certification")

    # Row 1: banner (two merged title labels at B and F)
    ws["B1"] = "Foundational Qualification: Education"
    ws["F1"] = "Foundational Qualification: Personal Certification"

    # Row 2: column headers
    ws["A2"] = "Work Role"
    ws["B2"], ws["C2"], ws["D2"] = "Basic", "Intermediate", "Advanced"
    ws["E2"] = "- OR -"
    ws["F2"], ws["G2"], ws["H2"] = "Basic", "Intermediate", "Advanced"

    # Data rows
    ws["A3"] = "(411) Technical Support Specialist"
    ws["B3"] = "BS in IT, Cyber, Data Sci, Info Sys, Comp Sci"
    ws["D3"] = "TBD"
    ws["F3"] = "A+, Network+"
    ws["G3"] = "GFACT, CND, Security+, GSEC"
    ws["H3"] = "FITSP-O, GICSP, CASP+, CCNP-Security, CISA, SSCP"

    ws["A4"] = "(421) Database Administrator"
    ws["B4"] = "BS in IT, Cyber, Data Sci, Info Sys, Comp Sci"
    ws["D4"] = "TBD"
    ws["G4"] = "Cloud+, SSCP, Security+, GSEC"
    ws["H4"] = "CASP+, CCNP-Security, CISA, CISSP, ISSAP, ISSEP"

    out = Path("scripts/tests/fixtures/tiny_jan2025.xlsx")
    out.parent.mkdir(exist_ok=True)
    wb.save(out)
    print(f"wrote {out}")


if __name__ == "__main__":
    build()
