"""Build a tiny fixture xlsx mimicking the V2.1 Certification Repository sheet.

Run once from repo root:
    .venv/Scripts/python -m scripts.tests.build_v21_fixture
"""
from pathlib import Path
from openpyxl import Workbook


def build() -> None:
    wb = Workbook()
    wb.remove(wb.active)
    ws = wb.create_sheet("Certification Repository")

    # Row 1: decorative header markers the real sheet has in non-data columns
    ws["C1"] = "Pending OPR Validation"
    ws["G1"] = "Added"
    ws["H1"] = "Modified"

    # Row 2: column headers
    ws.append(["WRC", "Work Role Title", "Element", "Acronym", "Proficiency", "Vendor"])

    # Row 3+: data covering edge cases:
    # - multiple certs at same level for same role
    # - same role at multiple levels
    # - trailing spaces in text (intentional; mirrors real file)
    # - distinct vendors
    data_rows = [
        (111, "All-Source Analyst ", "Intelligence (Cyberspace)", "RCCE Level 1", "Advanced", "Rocheston LLC"),
        (111, "All-Source Analyst ", "Intelligence (Cyberspace)", "CySA+", "Advanced", "CompTIA, Inc."),
        (211, "Forensics Analyst", "Cyberspace Enablers", "CHFI", "Intermediate", "EC-Council"),
        (211, "Forensics Analyst", "Cyberspace Enablers", "PenTest+", "Intermediate", "CompTIA, Inc."),
        (211, "Forensics Analyst", "Cyberspace Enablers", "CySA+", "Advanced", "CompTIA, Inc."),
        (211, "Forensics Analyst", "Cyberspace Enablers", "GCFA", "Advanced", "GIAC"),
        (411, "Technical Support Specialist", "IT (Cyberspace)", "A+", "Basic", "CompTIA, Inc."),
        (411, "Technical Support Specialist", "IT (Cyberspace)", "Network+", "Basic", "CompTIA, Inc."),
        (411, "Technical Support Specialist", "IT (Cyberspace)", "Security+", "Intermediate", "CompTIA, Inc."),
    ]
    for row in data_rows:
        ws.append(row)

    out = Path("scripts/tests/fixtures/tiny_v21.xlsx")
    out.parent.mkdir(exist_ok=True)
    wb.save(out)
    print(f"wrote {out}")


if __name__ == "__main__":
    build()
