"""Normalize the Jan 2025 'Education or Certification' sheet to canonical records.

The Jan 2025 workbook's layout (pre-V2.1 paradigm):
  Column A: role (formatted "(CODE) Name")
  Columns B/C/D: Education at Basic/Intermediate/Advanced
  Column E: " - OR - " literal
  Columns F/G/H: Certification at Basic/Intermediate/Advanced

Each cell is a comma-separated cert/education list. "TBD" and blanks → empty.
"""
import json
import re
import sys
from pathlib import Path

from openpyxl import load_workbook

from scripts.schema import MatrixRecord

ROLE_CODE_RE = re.compile(r"^\((\d+)\)\s*(.+)$")

# 0-indexed column mapping (iter_rows values_only tuple)
COL_ROLE = 0
COL_ED_BASIC = 1
COL_ED_INTER = 2
COL_ED_ADV = 3
# 4 = " - OR - "
COL_CERT_BASIC = 5
COL_CERT_INTER = 6
COL_CERT_ADV = 7


NO_CONTENT_MARKERS = {"", "TBD", "<BLANK>", "N/A", "NA", "-"}


def _split_certs(cell_value: object) -> list[str]:
    """Split a cert-list cell into a list of cert names.

    Treat as empty: None, blank strings, and explicit no-content markers
    like 'TBD', '<blank>', 'N/A' (Jeff's spreadsheet uses '<blank>' as an
    intentional "no certs at this level" marker — 40 cells).
    """
    if cell_value is None:
        return []
    s = str(cell_value).strip()
    if s.upper() in NO_CONTENT_MARKERS:
        return []
    return [c.strip() for c in s.split(",") if c.strip()]


def normalize_xlsx(path: str | Path) -> list[MatrixRecord]:
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb["Education or Certification"]
    records: list[MatrixRecord] = []
    for row in ws.iter_rows(min_row=3, values_only=True):
        if not row or not row[COL_ROLE]:
            continue
        m = ROLE_CODE_RE.match(str(row[COL_ROLE]).strip())
        if not m:
            continue
        code = m.group(1)
        name = m.group(2).strip()
        per_type_per_level = [
            ("education", "basic", _split_certs(row[COL_ED_BASIC])),
            ("education", "intermediate", _split_certs(row[COL_ED_INTER])),
            ("education", "advanced", _split_certs(row[COL_ED_ADV])),
            ("certification", "basic", _split_certs(row[COL_CERT_BASIC])),
            ("certification", "intermediate", _split_certs(row[COL_CERT_INTER])),
            ("certification", "advanced", _split_certs(row[COL_CERT_ADV])),
        ]
        for qtype, level, certs in per_type_per_level:
            records.append(MatrixRecord(
                work_role_code=code,
                work_role_name=name,
                qualification_type=qtype,
                proficiency_level=level,
                certs=certs,
            ))
    wb.close()
    return records


if __name__ == "__main__":
    if len(sys.argv) < 2:
        sys.stderr.write(
            "usage: python -m scripts.normalize_jan2025_xlsx PATH_TO_JAN2025_XLSX\n"
            "\n"
            "This script produces the informational CHANGELOG only. The author's\n"
            "private Jan 2025 working version is not committed to this repo; if\n"
            "you don't have access to it, you don't need to run this script —\n"
            "the published 8140/CHANGELOG.md is already generated.\n"
        )
        sys.exit(2)
    src = Path(sys.argv[1])
    out = Path("data/jan2025.json")
    out.parent.mkdir(exist_ok=True)
    records = normalize_xlsx(src)
    out.write_text(
        json.dumps([r.to_canonical() for r in records], indent=2),
        encoding="utf-8",
    )
    print(f"wrote {len(records)} records to {out} from {src}")
