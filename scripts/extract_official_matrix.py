"""Extract the current DoD 8140 qualification matrix from the V2.1 xlsx.

Source sheet: `Certification Repository` — a flat table with columns
    WRC | Work Role Title | Element | Acronym | Proficiency | Vendor

Each row represents one cert option for one (work role, proficiency level).
Rows are grouped by (work_role_code, proficiency_level) and emitted as
canonical MatrixRecords with `qualification_type="certification"`.

Out of scope: Education, DoD Training, Commercial Training paths. Those are
other sheets in the source workbook and are intentionally not extracted here
per the project design (see docs/refresh-notes-2026-04.md).
"""
import json
import sys
from collections import defaultdict
from pathlib import Path

from openpyxl import load_workbook

from scripts.schema import MatrixRecord

SHEET_NAME = "Certification Repository"
HEADER_ROW = 2  # Row 1 has decorative markers; row 2 is the column header
COL_WRC = 0
COL_ROLE_NAME = 1
COL_ACRONYM = 3
COL_PROFICIENCY = 4


def extract_from_xlsx(path: str | Path) -> list[MatrixRecord]:
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb[SHEET_NAME]

    # key = (role_code, proficiency_level_lower) → {"name": str, "certs": set[str]}
    buckets: dict[tuple[str, str], dict] = defaultdict(
        lambda: {"name": "", "certs": set()}
    )

    for row in ws.iter_rows(min_row=HEADER_ROW + 1, values_only=True):
        if row[COL_WRC] is None or row[COL_ACRONYM] is None:
            continue
        code = str(row[COL_WRC]).strip()
        name = str(row[COL_ROLE_NAME]).strip() if row[COL_ROLE_NAME] else ""
        acronym = str(row[COL_ACRONYM]).strip()
        proficiency = (
            str(row[COL_PROFICIENCY]).strip().lower()
            if row[COL_PROFICIENCY]
            else ""
        )
        if not code or not acronym or not proficiency:
            continue
        key = (code, proficiency)
        bucket = buckets[key]
        if not bucket["name"]:
            bucket["name"] = name
        bucket["certs"].add(acronym)

    wb.close()

    records = [
        MatrixRecord(
            work_role_code=code,
            work_role_name=bucket["name"],
            qualification_type="certification",
            proficiency_level=level,
            certs=sorted(bucket["certs"]),
        )
        for (code, level), bucket in sorted(buckets.items())
    ]
    return records


if __name__ == "__main__":
    src = (
        Path(sys.argv[1])
        if len(sys.argv) > 1
        else Path("8140/sources/dod8140-matrix-v2.1-20250919.xlsx")
    )
    out = Path("data/official.json")
    out.parent.mkdir(exist_ok=True)
    records = extract_from_xlsx(src)
    out.write_text(
        json.dumps([r.to_canonical() for r in records], indent=2),
        encoding="utf-8",
    )
    print(f"wrote {len(records)} records to {out} from {src}")
