"""Build the refreshed 8140 xlsx from V2.1 source data.

The Jan 2025 xlsx is Jeff's **visual template** — its tab layout and the
style of the pivot view we recreate. It is NOT patched; we build a fresh
workbook from V2.1's Certification Repository.

Output structure:
  - 'Explanation'              placeholder narrative (Jeff rewrites)
  - 'Certification Requirements' per-role cert list at each proficiency level
  - 'Certification Analysis'   inverted pivot: roles x certs, with 1/2/3 cells
                               + Total Positions / Total Points summary rows
"""
import sys
from collections import defaultdict
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.page import PageMargins

import colorsys

from scripts.visual_spec import (
    CERT_COLOR_OVERRIDES,
    CERT_ORDER_BY_VENDOR,
    CERT_SHORT_NAMES,
    CY101_SEPARATOR_BEFORE_CODE,
    DEFAULT_PALETTE,
    DEFAULT_VENDOR_HEADER_FONT_SIZE,
    ROLE_NAME_OVERRIDES,
    ROLE_ORDER,
    ROLE_ROW_HIGHLIGHTS,
    VENDOR_HEADER_FONT_SIZE,
    VENDOR_HUE_SPEC,
    VENDOR_ORDER,
    VENDOR_PALETTE,
    VENDOR_SHORT_NAMES,
)


# Lightness values for proficiency levels. Lower = darker.
# Deliberate wide spread so Basic vs Advanced reads at a glance.
LEVEL_LIGHTNESS = {1: 0.88, 2: 0.60, 3: 0.25}
HEADER_CELL_LIGHTNESS = 0.18  # cert acronym header — darker than Advanced


def _hsl_to_argb(hue_deg: float, sat: float, lightness: float) -> str:
    """Convert HSL (hue in degrees, sat+lightness 0-1) to Excel ARGB string."""
    h = (hue_deg % 360) / 360.0
    r, g, b = colorsys.hls_to_rgb(h, lightness, sat)
    return f"FF{int(r * 255):02X}{int(g * 255):02X}{int(b * 255):02X}"


def _cert_color_spec(short_cert: str, vendor_short: str,
                     cert_index: int, total_certs: int) -> tuple[float, float]:
    """Return (hue_deg, saturation) for a cert column. Overrides win."""
    if short_cert in CERT_COLOR_OVERRIDES:
        return CERT_COLOR_OVERRIDES[short_cert]
    spec = VENDOR_HUE_SPEC.get(vendor_short)
    if spec is None:
        return 0.0, 0.0
    if total_certs <= 1:
        t = 0.0
    else:
        t = cert_index / (total_certs - 1)
    hue = spec["hue_start"] + t * (spec["hue_end"] - spec["hue_start"])
    return hue, spec["sat"]


def _cert_column_color(short_cert: str, vendor_short: str,
                        cert_index: int, total_certs: int) -> tuple[str, str]:
    """Return (cell_fill_argb, text_color_argb) for a cert column.

    One color per cert column, applied to both the cert acronym header and
    every data cell in that column. Level (1/2/3) is communicated by the
    cell VALUE, not the fill. Within a vendor group the lightness walks
    wider (easier=lighter, harder=darker) so adjacent cert columns are
    visually distinguishable even inside single-hue vendors like CompTIA.

    Text color auto-picks white on dark fills, black on light fills."""
    hue, sat = _cert_color_spec(short_cert, vendor_short, cert_index, total_certs)
    if total_certs <= 1:
        lightness = 0.50
    else:
        t = cert_index / (total_certs - 1)
        lightness = 0.78 - t * 0.58  # 0.78 -> 0.20
    fill = _hsl_to_argb(hue, sat, lightness)
    text = "FFFFFFFF" if lightness < 0.55 else "FF000000"
    return fill, text


def _cert_header_fill(short_cert: str, vendor_short: str, cert_index: int,
                      total_certs: int) -> tuple[str, str]:
    """Header variant of the cert column color — same hue, but always
    toward the darker end of the walk so the header stands out above the
    (lighter) cells."""
    hue, sat = _cert_color_spec(short_cert, vendor_short, cert_index, total_certs)
    if total_certs <= 1:
        lightness = 0.28
    else:
        t = cert_index / (total_certs - 1)
        lightness = 0.42 - t * 0.28  # 0.42 -> 0.14
    fill = _hsl_to_argb(hue, sat, lightness)
    text = "FFFFFFFF" if lightness < 0.55 else "FF000000"
    return fill, text


def _heatmap_fill(value: int, max_value: int, kind: str) -> str | None:
    """Return fill color for a heat-map cell. kind=positions or points."""
    if value <= 0 or max_value <= 0:
        return None
    t = value / max_value
    # Two palettes: positions (green-red scale) and points (similar).
    # Both use lightness driven by t.
    if kind == "positions":
        # light pink -> dark red scale, per Jan 2025 convention
        return _hsl_to_argb(hue_deg=0, sat=0.55, lightness=0.90 - 0.55 * t)
    # points
    return _hsl_to_argb(hue_deg=130, sat=0.45, lightness=0.90 - 0.55 * t)

SHEET_NAME = "Certification Repository"
HEADER_ROW = 2
COL_WRC = 0
COL_ROLE_NAME = 1
COL_ELEMENT = 2
COL_ACRONYM = 3
COL_PROFICIENCY = 4
COL_VENDOR = 5

PROFICIENCY_LEVEL = {"basic": 1, "intermediate": 2, "advanced": 3}
LEVEL_LABEL = {1: "Basic", 2: "Intermediate", 3: "Advanced"}

# Work roles in V2.1 role universe but with no published cert options.
# Surfaced as a footnote/note in the output xlsx rather than matrix rows.
PENDING_REVIEW_ROLES = {
    "462": "Control Systems Security Specialist",
    "731": "Cyber Legal Advisor",
    "901": "Executive Cyber Leader",
}


# ----------------------------------------------------------------------------
# Data extraction helpers
# ----------------------------------------------------------------------------

def read_v21_certification_rows(xlsx_path: str | Path) -> list[dict]:
    """Read the V2.1 Certification Repository sheet into a list of dicts."""
    wb = load_workbook(xlsx_path, read_only=True, data_only=True)
    ws = wb[SHEET_NAME]
    rows: list[dict] = []
    for row in ws.iter_rows(min_row=HEADER_ROW + 1, values_only=True):
        if row[COL_WRC] is None or row[COL_ACRONYM] is None:
            continue
        proficiency = (
            str(row[COL_PROFICIENCY]).strip().lower()
            if row[COL_PROFICIENCY] else ""
        )
        if proficiency not in PROFICIENCY_LEVEL:
            continue
        rows.append({
            "wrc": str(row[COL_WRC]).strip(),
            "work_role_name": str(row[COL_ROLE_NAME]).strip() if row[COL_ROLE_NAME] else "",
            "element": str(row[COL_ELEMENT]).strip() if row[COL_ELEMENT] else "",
            "acronym": str(row[COL_ACRONYM]).strip(),
            "proficiency": proficiency,
            "vendor": str(row[COL_VENDOR]).strip() if row[COL_VENDOR] else "",
        })
    wb.close()
    return rows


def build_role_catalog(rows: list[dict]) -> dict[str, dict]:
    """role_code -> {name, element}, deduped."""
    catalog: dict[str, dict] = {}
    for r in rows:
        code = r["wrc"]
        if code not in catalog:
            catalog[code] = {"name": r["work_role_name"], "element": r["element"]}
    return catalog


def build_vendor_cert_map(rows: list[dict]) -> dict[str, list[str]]:
    """vendor -> sorted list of cert acronyms."""
    vendor_sets: dict[str, set] = defaultdict(set)
    for r in rows:
        vendor_sets[r["vendor"]].add(r["acronym"])
    return {v: sorted(certs) for v, certs in vendor_sets.items()}


def build_pivot_cells(rows: list[dict]) -> dict[tuple[str, str], int]:
    """(role_code, cert_acronym) -> highest proficiency_level as int 1/2/3."""
    cells: dict[tuple[str, str], int] = {}
    for r in rows:
        key = (r["wrc"], r["acronym"])
        level = PROFICIENCY_LEVEL[r["proficiency"]]
        if key not in cells or cells[key] < level:
            cells[key] = level
    return cells


# ----------------------------------------------------------------------------
# Sheet writers
# ----------------------------------------------------------------------------

HEADER_FILL = PatternFill("solid", start_color="FF1F4E79", end_color="FF1F4E79")
HEADER_FONT = Font(bold=True, color="FFFFFFFF")
VENDOR_GROUP_FILL = PatternFill("solid", start_color="FFDCE6F1", end_color="FFDCE6F1")
VENDOR_GROUP_FONT = Font(bold=True)
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT_WRAP = Alignment(horizontal="left", vertical="center", wrap_text=True)


# Default Explanation block used on the FIRST build of an output xlsx.
# On subsequent rebuilds, if the output file already exists, the builder
# preserves whatever text Jeff has typed into the rows below the
# pending-review footnote (the footnote text serves as the sentinel).
# Jeff edits the narrative directly in Excel — no code changes needed.
DEFAULT_EXPLANATION_LINES: list[str] = [
    "",  # spacer
    "DoD 8140 Cybersecurity Workforce Qualification — Personal Certification Path",
    (
        "Governed by DoDM 8140.03 \"Cyberspace Workforce Qualification and Management Program\" (15 Feb 2023) — "
        "https://dodcio.defense.gov/Portals/0/Documents/Library/DoDM-8140-03.pdf. "
        "This matrix is compiled from the DoD 8140 Foundational Qualification Matrix V2.1 (effective 19 Sep 2025), "
        "published at https://www.cyber.mil/dod-workforce-innovation-directorate/dod8140/qualification-matrices."
    ),
    (
        "Work roles: every IT and cyber-related DoD position is mapped to one or more \"work roles\" — "
        "about 70 roles grouped into 7 workforce elements (Cybersecurity, IT, Cyber Enablers, Cyber Effects, "
        "Intelligence, Software Engineering, Data/AI). A single individual may hold up to 3 work roles; qualification "
        "is evaluated per role. See the DCWF at https://www.cyber.mil/dod-workforce-innovation-directorate/dod-cyber-workforce-framework. "
        "Each role has three proficiency levels (Basic, Intermediate, Advanced) — a higher-level qualification also satisfies "
        "lower levels within the same role."
    ),
    (
        "Qualification paths: 8140 qualification can be met via one of four \"Foundational Qualification\" paths — "
        "Education, DoD/Military Training, Commercial Training, or Personnel Certification — plus an \"Experience\" "
        "alternative for eligible personnel. This matrix shows only the Personnel Certification path, which may be "
        "the most practical route for anyone without an applicable degree or military schooling."
    ),
    (
        "Education path: Bachelor's, Master's, or PhD in Information Technology, Cybersecurity, Data Science, "
        "Information Systems, Computer Science, or Software Engineering from an ABET-accredited (abet.org) or "
        "CAE-designated (caecommunity.org) institution. Bachelor's satisfies Basic and Intermediate levels; "
        "Master's or PhD satisfies all three levels, including Advanced (a V2.1 change — previously TBD at Advanced). "
        "Degree must be within the last 5 years with documented continuous cyberspace work since. See the "
        "\"8140 Interim Education Qualification Options\" reference cited throughout the DoD matrix workbook for "
        "additional considerations."
    ),
    (
        "DoD / Military Training path: specific military schoolhouse courses listed per work role in the DoD 8140 "
        "Training Repository (within the same workbook linked above). Generally not accessible to government civilians "
        "or contractors unless previously completed during military service. Exception: CY 101 (40-hour online course) "
        "satisfies qualification for all Cyber Enabler work roles — https://www.cyber.mil/training/cyber-101."
    ),
    (
        "Commercial Training path: commercial courses approved as 8140 qualification options. "
        "Thin option universe; most entries overlap with certifications."
    ),
    (
        "Experience (alternative): available only to government civilians and military personnel. "
        "Contractors are NOT eligible for the Experience path."
    ),
    "Compiled by Jeff Krueger. Unaffiliated with any firm, agency, or contracting office. Not legal advice.",
]

# Sentinel pattern in the footnote row — used to locate the boundary
# between generated matrix content and user-editable narrative.
FOOTNOTE_SENTINEL = "pending DoD review"


def _read_existing_narrative(output_path: str | Path) -> list[str] | None:
    """If an output xlsx already exists, return the narrative lines the user
    has authored below the pending-review footnote. Returns None if the file
    is absent or unreadable (first build, or sheet structure changed) — the
    caller falls back to DEFAULT_EXPLANATION_LINES in that case."""
    p = Path(output_path)
    if not p.exists():
        return None
    try:
        wb = load_workbook(p, data_only=True)
    except Exception:
        return None
    if "Certification Analysis" not in wb.sheetnames:
        wb.close()
        return None
    ws = wb["Certification Analysis"]
    footnote_row = None
    for r in range(1, ws.max_row + 1):
        v = ws.cell(row=r, column=1).value
        if v and FOOTNOTE_SENTINEL in str(v):
            footnote_row = r
            break
    if footnote_row is None:
        wb.close()
        return None
    narrative: list[str] = []
    for r in range(footnote_row + 1, ws.max_row + 1):
        v = ws.cell(row=r, column=1).value
        narrative.append(str(v) if v is not None else "")
    wb.close()
    # Trim trailing empties so we don't grow the output file by a blank per build.
    while narrative and not narrative[-1]:
        narrative.pop()
    return narrative if narrative else None


def _vendor_short_name(v: str) -> str:
    return VENDOR_SHORT_NAMES.get(v, v)


def _cert_short_name(c: str) -> str:
    return CERT_SHORT_NAMES.get(c, c)


def _build_cert_column_layout(rows: list[dict]) -> list[tuple[str, str]]:
    """Return ordered list of (short_cert, short_vendor) tuples for pivot columns.

    Within each vendor group, certs are sorted by **average proficiency level
    across all roles that require them**, ascending. This puts entry-level
    certs (mostly Basic, avg near 1) before advanced certs (mostly Advanced,
    avg near 3) — loosely in order of difficulty. Ties broken by cert name.
    """
    # Index V2.1 data by short vendor, collecting cert acronym and proficiency.
    by_vendor: dict[str, set[str]] = {}
    cert_levels: dict[str, list[int]] = {}
    for r in rows:
        vs = _vendor_short_name(r["vendor"])
        cs = _cert_short_name(r["acronym"])
        level = PROFICIENCY_LEVEL.get(r["proficiency"])
        if level is None:
            continue
        by_vendor.setdefault(vs, set()).add(cs)
        cert_levels.setdefault(cs, []).append(level)

    def avg_level(cert: str) -> float:
        levels = cert_levels.get(cert, [])
        return sum(levels) / len(levels) if levels else 0.0

    layout: list[tuple[str, str]] = []
    known_vendors: list[str] = list(VENDOR_ORDER)
    for v in known_vendors:
        certs = by_vendor.get(v)
        if not certs:
            continue
        ordered = sorted(certs, key=lambda c: (avg_level(c), c.lower()))
        for cert in ordered:
            layout.append((cert, v))

    # Unknown vendors (not in VENDOR_ORDER), appended at end; same sort rule.
    for v in sorted(set(by_vendor) - set(known_vendors)):
        ordered = sorted(by_vendor[v], key=lambda c: (avg_level(c), c.lower()))
        for cert in ordered:
            layout.append((cert, v))

    return layout


def _build_pivot_cells_short(rows: list[dict]) -> dict[tuple[str, str], int]:
    """(role_code, short_cert_name) -> highest proficiency level 1/2/3."""
    cells: dict[tuple[str, str], int] = {}
    for r in rows:
        key = (r["wrc"], _cert_short_name(r["acronym"]))
        level = PROFICIENCY_LEVEL[r["proficiency"]]
        if key not in cells or cells[key] < level:
            cells[key] = level
    return cells


def _build_role_row_order(role_catalog: dict) -> list[str]:
    """Sorted ascending by WRC. Pending-review roles omitted. No separator
    row (CY 101 guidance doesn't cleanly map to pure numerical ordering, so
    we leave it out entirely per Jeff's direction)."""
    excluded = set(PENDING_REVIEW_ROLES)
    return sorted(
        (c for c in role_catalog if c not in excluded),
        key=lambda c: int(c),
    )


# Cell alignment / style presets
CERT_HEADER_ROT = Alignment(horizontal="center", vertical="bottom", text_rotation=90, wrap_text=False)
CELL_CENTER = Alignment(horizontal="center", vertical="center")


def _outline_box(ws, top_row: int, left_col: int,
                 bottom_row: int, right_col: int) -> None:
    """Draw a thin outer border around a rectangular range, preserving any
    existing borders on interior cells."""
    thin = Side(style="thin", color="FF000000")
    # Top edge
    for c in range(left_col, right_col + 1):
        cell = ws.cell(row=top_row, column=c)
        b = cell.border
        cell.border = Border(
            top=thin,
            left=thin if c == left_col else b.left,
            right=thin if c == right_col else b.right,
            bottom=b.bottom,
        )
    # Bottom edge
    for c in range(left_col, right_col + 1):
        cell = ws.cell(row=bottom_row, column=c)
        b = cell.border
        cell.border = Border(
            bottom=thin,
            left=thin if c == left_col else b.left,
            right=thin if c == right_col else b.right,
            top=b.top,
        )
    # Left and right edges (excluding corners already handled above)
    for r in range(top_row + 1, bottom_row):
        left = ws.cell(row=r, column=left_col)
        b = left.border
        left.border = Border(left=thin, top=b.top, right=b.right, bottom=b.bottom)
        if right_col != left_col:
            right = ws.cell(row=r, column=right_col)
            b = right.border
            right.border = Border(right=thin, top=b.top, left=b.left, bottom=b.bottom)


def _palette_for(vendor_short: str) -> dict:
    return VENDOR_PALETTE.get(vendor_short, DEFAULT_PALETTE)


def _cert_index_within_vendor(cert_columns: list[tuple[str, str]]) -> dict[int, tuple[int, int]]:
    """Map global column index -> (index-within-vendor, total-certs-for-vendor)."""
    by_vendor_positions: dict[str, list[int]] = {}
    for idx, (_, vendor) in enumerate(cert_columns):
        by_vendor_positions.setdefault(vendor, []).append(idx)
    mapping: dict[int, tuple[int, int]] = {}
    for vendor, positions in by_vendor_positions.items():
        total = len(positions)
        for within_idx, global_idx in enumerate(positions):
            mapping[global_idx] = (within_idx, total)
    return mapping


def write_pivot_sheet(
    wb: Workbook,
    role_catalog: dict,
    rows: list[dict],
    narrative_lines: list[str] | None = None,
) -> None:
    ws = wb.create_sheet("Certification Analysis")

    cert_columns = _build_cert_column_layout(rows)
    pivot_cells = _build_pivot_cells_short(rows)
    role_rows = _build_role_row_order(role_catalog)
    cert_pos = _cert_index_within_vendor(cert_columns)

    first_cert_col = 2  # Column B
    last_cert_col = first_cert_col + len(cert_columns) - 1
    echo_col = last_cert_col + 1  # right-hand Work Role echo column

    banner_fill = PatternFill("solid", fgColor="FF1F4E79")
    white_bold = Font(bold=True, color="FFFFFFFF")
    white_bold_small = Font(bold=True, color="FFFFFFFF", size=10)

    # ----- Row 1: banner -----
    ws.cell(row=1, column=1, value="DoD 8140.03 Foundational Qualification: Personal Certification").font = Font(
        bold=True, size=12, color="FFFFFFFF"
    )
    ws.cell(row=1, column=1).fill = banner_fill
    ws.cell(row=1, column=1).alignment = CELL_CENTER
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=echo_col)

    # ----- Row 2: vendor group headers (merged per vendor) -----
    col_cursor = first_cert_col
    vendor_iter = list(VENDOR_ORDER) + sorted(
        {vend for _, vend in cert_columns if vend not in VENDOR_ORDER}
    )
    vendor_spans: list[tuple[int, int]] = []
    for v in vendor_iter:
        certs_for_v = [c for c, vv in cert_columns if vv == v]
        if not certs_for_v:
            continue
        span = len(certs_for_v)
        start = col_cursor
        end = col_cursor + span - 1
        pal = _palette_for(v)
        cell = ws.cell(row=2, column=start, value=v)
        font_size = VENDOR_HEADER_FONT_SIZE.get(v, DEFAULT_VENDOR_HEADER_FONT_SIZE)
        cell.font = Font(bold=True, color="FFFFFFFF", size=font_size)
        cell.fill = PatternFill("solid", fgColor=pal["base"])
        cell.alignment = CELL_CENTER
        if span > 1:
            ws.merge_cells(start_row=2, start_column=start, end_row=2, end_column=end)
        vendor_spans.append((start, end))
        col_cursor = end + 1

    # ----- Row 3: cert acronym headers (each cert its own color) -----
    for c in (1, echo_col):
        hcell = ws.cell(row=3, column=c, value="Work Role")
        hcell.font = white_bold
        hcell.fill = banner_fill
        hcell.alignment = CELL_CENTER
    for i, (cert, vendor) in enumerate(cert_columns):
        within_idx, total = cert_pos[i]
        header_fill, header_text = _cert_header_fill(cert, vendor, within_idx, total)
        cell = ws.cell(row=3, column=first_cert_col + i, value=cert)
        cell.font = Font(bold=True, color=header_text, size=10)
        cell.fill = PatternFill("solid", fgColor=header_fill)
        cell.alignment = CERT_HEADER_ROT

    # ----- Role rows (pure numerical; no CY 101 separator) -----
    band_fill = PatternFill("solid", fgColor="FFF2F2F2")  # very light gray banding

    current_row = 4
    first_data_row = 4
    last_data_row = 4
    data_row_count = 0
    for code in role_rows:
        name = ROLE_NAME_OVERRIDES.get(code, role_catalog[code]["name"])
        role_label = f"({code}) {name}"
        band = (data_row_count % 2 == 1)
        data_row_count += 1
        a_cell = ws.cell(row=current_row, column=1, value=role_label)
        a_cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        echo_cell = ws.cell(row=current_row, column=echo_col, value=role_label)
        echo_cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        highlight = ROLE_ROW_HIGHLIGHTS.get(code)
        if highlight:
            fill = PatternFill("solid", fgColor=highlight)
            a_cell.fill = fill
            a_cell.font = Font(bold=True, color="FFFFFFFF")
            echo_cell.fill = fill
            echo_cell.font = Font(bold=True, color="FFFFFFFF")
        elif band:
            a_cell.fill = band_fill
            echo_cell.fill = band_fill
        for i, (cert, vendor) in enumerate(cert_columns):
            level = pivot_cells.get((code, cert))
            if level is None:
                if band:
                    ws.cell(row=current_row, column=first_cert_col + i).fill = band_fill
                continue
            within_idx, total = cert_pos[i]
            cell_fill, text_color = _cert_column_color(cert, vendor, within_idx, total)
            cell = ws.cell(row=current_row, column=first_cert_col + i, value=level)
            cell.alignment = CELL_CENTER
            cell.fill = PatternFill("solid", fgColor=cell_fill)
            cell.font = Font(bold=True, color=text_color)
        last_data_row = current_row
        current_row += 1

    # ----- Combined legend + repeated cert header row (no blank spacer) -----
    # Column A: proficiency legend.
    # Cert columns: repeated cert acronym headers, rotated 90, per-cert shading.
    # Echo column: the "darker shades cover more work roles" note (replaces
    # a separate margin column).
    repeat_header_row = current_row
    legend_cell = ws.cell(
        row=repeat_header_row, column=1,
        value="Proficiency levels:\nBasic = 1, Intermediate = 2, Advanced = 3",
    )
    legend_cell.font = Font(italic=True, size=9)
    legend_cell.alignment = Alignment(horizontal="left", vertical="bottom", wrap_text=True)

    echo_note = ws.cell(
        row=repeat_header_row, column=echo_col,
        value="darker shades cover more work roles",
    )
    echo_note.font = Font(italic=True, size=9)
    echo_note.alignment = Alignment(horizontal="left", vertical="bottom", wrap_text=True)

    for i, (cert, vendor) in enumerate(cert_columns):
        within_idx, total = cert_pos[i]
        header_fill, header_text = _cert_header_fill(cert, vendor, within_idx, total)
        cell = ws.cell(row=repeat_header_row, column=first_cert_col + i, value=cert)
        cell.font = Font(bold=True, color=header_text, size=10)
        cell.fill = PatternFill("solid", fgColor=header_fill)
        cell.alignment = CERT_HEADER_ROT
    ws.row_dimensions[repeat_header_row].height = 40
    current_row += 1

    # ----- Thin outer box around each vendor's column group (row 2 -> repeat header) -----
    for start_col, end_col in vendor_spans:
        _outline_box(ws, top_row=2, left_col=start_col,
                     bottom_row=repeat_header_row, right_col=end_col)

    # ----- Summary rows (formulas + heatmap fills) -----
    totals_row = current_row
    ws.cell(row=totals_row, column=1, value="Total Positions Covered").font = Font(bold=True)
    ws.cell(row=totals_row, column=echo_col, value="Total Positions Covered").font = Font(bold=True)
    current_row += 1

    points_row = current_row
    ws.cell(row=points_row, column=1, value='Total "Points" (proficiency levels \u00d7 positions)').font = Font(bold=True)
    ws.cell(row=points_row, column=echo_col, value='Total "Points"').font = Font(bold=True)
    current_row += 1

    # Pre-compute max values per row for heat-map scaling
    positions_values: list[int] = []
    points_values: list[int] = []
    for i, (cert, _vendor) in enumerate(cert_columns):
        positions_values.append(
            sum(1 for code in role_catalog if (code, cert) in pivot_cells)
        )
        points_values.append(
            sum(pivot_cells[(code, cert)] for code in role_catalog if (code, cert) in pivot_cells)
        )
    max_positions = max(positions_values) if positions_values else 0
    max_points = max(points_values) if points_values else 0

    for i in range(len(cert_columns)):
        col_letter = get_column_letter(first_cert_col + i)
        data_range = f"{col_letter}{first_data_row}:{col_letter}{last_data_row}"
        # Formulas (auditable)
        tcell = ws.cell(row=totals_row, column=first_cert_col + i, value=f"=COUNT({data_range})")
        pcell = ws.cell(row=points_row, column=first_cert_col + i, value=f"=SUM({data_range})")
        tcell.alignment = CELL_CENTER
        pcell.alignment = CELL_CENTER
        # Heatmap fills driven by pre-computed values
        pos_fill = _heatmap_fill(positions_values[i], max_positions, "positions")
        pts_fill = _heatmap_fill(points_values[i], max_points, "points")
        if pos_fill:
            tcell.fill = PatternFill("solid", fgColor=pos_fill)
        if pts_fill:
            pcell.fill = PatternFill("solid", fgColor=pts_fill)
        # Bold text on the high-coverage cells
        high_threshold_pos = 0.7 * max_positions if max_positions else 0
        high_threshold_pts = 0.7 * max_points if max_points else 0
        if positions_values[i] >= high_threshold_pos and positions_values[i] > 0:
            tcell.font = Font(bold=True)
        if points_values[i] >= high_threshold_pts and points_values[i] > 0:
            pcell.font = Font(bold=True)

    # Explicit row heights on the summary block (v4 left these blank,
    # so Excel inherited the taller repeat-header height).
    ws.row_dimensions[totals_row].height = 14
    ws.row_dimensions[points_row].height = 14

    # ----- Outer box around the whole structured matrix (A1 through
    # last cert-value row in the echo column). Drawn BEFORE the footnote
    # so the footnote remains outside the frame.
    _outline_box(ws, top_row=1, left_col=1,
                 bottom_row=points_row, right_col=echo_col)

    # ----- Footnote (pending-review roles) — no blank spacer above -----
    footnote_row = current_row
    footnote = (
        "Note: The following work roles exist in DoD 8140 V2.1 but have no "
        "published certification options (pending DoD review): "
        + ", ".join(f"({c}) {n}" for c, n in PENDING_REVIEW_ROLES.items())
        + "."
    )
    fn_cell = ws.cell(row=footnote_row, column=1, value=footnote)
    fn_cell.font = Font(italic=True, color="FF595959", size=9)
    fn_cell.alignment = Alignment(wrap_text=True, vertical="center")
    ws.merge_cells(start_row=footnote_row, start_column=1, end_row=footnote_row, end_column=echo_col)
    ws.row_dimensions[footnote_row].height = 26

    # ----- Explanation block (below footnote, same printed page) -----
    # Narrative lines come either from an existing xlsx (preserving Jeff's
    # direct edits in Excel) or from DEFAULT_EXPLANATION_LINES on first build.
    explanation_lines = (
        narrative_lines if narrative_lines is not None else DEFAULT_EXPLANATION_LINES
    )
    explain_start_row = footnote_row + 1
    first_nonempty_seen = False
    for i, line in enumerate(explanation_lines):
        r = explain_start_row + i
        cell = ws.cell(row=r, column=1, value=line if line else None)
        if not line:
            ws.row_dimensions[r].height = 12  # thin spacer (2x prior)
            continue
        # First non-empty line of the narrative is styled as the title.
        is_title = not first_nonempty_seen
        first_nonempty_seen = True
        cell.font = Font(
            bold=is_title,
            italic=not is_title,
            size=20 if is_title else 18,  # ~2x the earlier 10/9 sizing
            color="FF000000" if is_title else "FF333333",
        )
        cell.alignment = Alignment(
            horizontal="left", vertical="top", wrap_text=True,
        )
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=echo_col)
        ws.row_dimensions[r].height = 28 if is_title else 24
    explain_end_row = explain_start_row + len(explanation_lines) - 1

    # ----- Column widths -----
    ws.column_dimensions["A"].width = 47
    for i in range(len(cert_columns)):
        letter = get_column_letter(first_cert_col + i)
        ws.column_dimensions[letter].width = 3.5
    ws.column_dimensions[get_column_letter(echo_col)].width = 47

    # ----- Row heights — squished ~10% from v4 baseline -----
    ws.row_dimensions[1].height = 17
    ws.row_dimensions[2].height = 17
    ws.row_dimensions[3].height = 40  # rotated cert acronyms (DAWIA prefix gone)
    for r in range(4, last_data_row + 1):
        if r not in ws.row_dimensions or ws.row_dimensions[r].height is None:
            ws.row_dimensions[r].height = 14

    # ----- Freeze panes: left of B, below row 3 -----
    ws.freeze_panes = "B4"

    # ----- Page setup for PDF export: Tabloid landscape, tight margins,
    # fit-to-1-page. Print area includes the matrix + footnote.
    ws.page_setup.orientation = "landscape"
    ws.page_setup.paperSize = 3  # Tabloid (11 x 17 in)
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 1
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_margins = PageMargins(
        left=0.25, right=0.25, top=0.25, bottom=0.25,
        header=0.1, footer=0.1,
    )
    ws.print_area = f"A1:{get_column_letter(echo_col)}{explain_end_row}"
    # Center on page horizontally (matrix is wider than tall).
    ws.print_options.horizontalCentered = True


# ----------------------------------------------------------------------------
# Entry point
# ----------------------------------------------------------------------------

def build(v21_xlsx_path: str | Path, output_path: str | Path) -> None:
    rows = read_v21_certification_rows(v21_xlsx_path)
    role_catalog = build_role_catalog(rows)

    # Preserve Jeff's narrative edits from a prior output, if present.
    # First build (or after structure change) falls back to defaults.
    preserved_narrative = _read_existing_narrative(output_path)

    wb = Workbook()
    wb.remove(wb.active)
    # Single consolidated sheet: matrix + explanation below it = 1-page PDF.
    write_pivot_sheet(wb, role_catalog, rows, narrative_lines=preserved_narrative)
    wb.properties.creator = "Jeff Krueger"
    wb.properties.lastModifiedBy = "Jeff Krueger"
    wb.properties.title = "DoD 8140.03 Cert Path Reference"
    wb.properties.description = (
        "DoD 8140.03 cybersecurity workforce qualification — certification-path "
        "reference. Compiled by Jeff Krueger. Unaffiliated with any firm or agency."
    )
    wb.save(output_path)


if __name__ == "__main__":
    src = (
        Path(sys.argv[1])
        if len(sys.argv) > 1
        else Path("8140/sources/dod8140-matrix-v2.1-20250919.xlsx")
    )
    out = (
        Path(sys.argv[2])
        if len(sys.argv) > 2
        else Path("8140/8140-cert-requirements.xlsx")
    )
    build(src, out)
    print(f"wrote {out}")
